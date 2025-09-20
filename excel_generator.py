import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
from config import DEFAULT_OUTPUT_EXCEL, logger

def create_excel_with_checkmarks(results, output_file=DEFAULT_OUTPUT_EXCEL):
    """
    Membuat file Excel dengan checklist untuk hasil analisis
    
    Parameters:
    results (list): Hasil analisis outlet
    output_file (str): Nama file output
    
    Returns:
    str: Path ke file output
    """
    try:
        # Konversi ke DataFrame pandas
        df = pd.DataFrame(results)
        
        # Simpan ke Excel
        df.to_excel(output_file, index=False)
        
        # Buka workbook yang sudah dibuat untuk ditambahkan formatting
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Styling untuk header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Styling untuk kolom True/False
        true_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        false_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Apply header styling
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Konversi True/False menjadi centang/silang dengan warna
        for row in range(2, ws.max_row + 1):
            for col in range(5, ws.max_column + 1):  # Kolom 5 dan seterusnya adalah kategori Boolean (setelah nama, koordinat, lat, lon)
                cell = ws.cell(row=row, column=col)
                if cell.value == True:
                    cell.value = "✓"
                    cell.fill = true_fill
                elif cell.value == False:
                    cell.value = "✗"
                    cell.fill = false_fill
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Simpan workbook yang sudah diformat
        wb.save(output_file)
        
        logger.info(f"File Excel berhasil dibuat: {output_file}")
        return os.path.abspath(output_file)
    
    except Exception as e:
        logger.error(f"Error saat membuat file Excel: {e}")
        return None

def add_summary_sheet(excel_file, summary_data):
    """
    Menambahkan lembar ringkasan ke file Excel yang sudah ada
    
    Parameters:
    excel_file (str): Path ke file Excel
    summary_data (dict): Data ringkasan untuk ditambahkan
    
    Returns:
    bool: True jika berhasil, False jika gagal
    """
    try:
        # Buka workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Buat sheet baru untuk ringkasan
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            # Hapus konten yang ada
            for row in ws.rows:
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet("Summary", 0)  # 0 untuk menempatkannya di awal
        
        # Styling
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font_color = Font(bold=True, color="FFFFFF")
        
        # Judul ringkasan
        ws['A1'] = "RINGKASAN ANALISIS OUTLET"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # Informasi jumlah total outlet
        ws['A3'] = "Total Outlet:"
        ws['A3'].font = header_font
        ws['B3'] = summary_data['total_outlets']
        
        # Header statistik kategori
        ws['A5'] = "Kategori"
        ws['A5'].font = header_font_color
        ws['A5'].fill = header_fill
        
        ws['B5'] = "Jumlah Outlet"
        ws['B5'].font = header_font_color
        ws['B5'].fill = header_fill
        
        ws['C5'] = "Persentase"
        ws['C5'].font = header_font_color
        ws['C5'].fill = header_fill
        
        # Statistik per kategori
        row = 6
        for category, (count, percentage) in summary_data['category_stats'].items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Informasi outlet dengan fasilitas terbanyak
        row += 2
        ws[f'A{row}'] = f"Outlet dengan fasilitas terbanyak ({summary_data['max_facilities']} fasilitas):"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:E{row}')
        
        # Daftar outlet terbaik
        for best_outlet in summary_data['best_outlets'][:10]:  # Batasi 10 outlet terbaik
            row += 1
            ws[f'A{row}'] = f"- {best_outlet}"
            ws.merge_cells(f'A{row}:E{row}')
        
        # Jika ada lebih dari 10 outlet terbaik
        if len(summary_data['best_outlets']) > 10:
            row += 1
            ws[f'A{row}'] = f"... dan {len(summary_data['best_outlets']) - 10} outlet lainnya"
            ws.merge_cells(f'A{row}:E{row}')
        
        # Atur lebar kolom
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Simpan perubahan
        wb.save(excel_file)
        
        logger.info(f"Berhasil menambahkan lembar ringkasan ke {excel_file}")
        return True
    
    except Exception as e:
        logger.error(f"Error saat menambahkan lembar ringkasan: {e}")
        return False